# Monthly Fulfillment Billing
import openpyxl

import pyautogui as pag

from openpyxl import load_workbook, workbook

# Set workbook and worksheet as variables
wb = load_workbook('MF.xlsx')
ws = wb.active

# Set Counter
i = 0

pag.moveTo(2017,187)
pag.PAUSE = .25
pag.mouseDown()
pag.PAUSE = .25
pag.mouseUp()

# Loop
for x in ws.rows:
    i = (i + 1)

        # Open Line Items tab
        #work
    pag.moveTo(1827,252)
    pag.click()


        # Add new Line
        #home
        #pag.moveTo(1286,336)
        #Office
    pag.moveTo(2038,379)
    pag.click(button = 'right')
    pag.PAUSE = .25
    pag.press('down')
    pag.PAUSE = .25
    pag.press('down')
    pag.PAUSE = .25
    pag.press('enter')
    pag.PAUSE = .25
        # Set variables

    item = ws.cell(row = i, column = 4).value
    pag.PAUSE = .25
    quantity = ws.cell(row = i, column = 6).value
    pag.PAUSE = .25
    price = ws.cell(row = i, column = 7).value
    pag.PAUSE = .25
    cost = ws.cell(row = i, column = 8).value
    pag.PAUSE = .25

        # Initial line data

        # Going to split the item into seperate key strokes
        #pag.press('capslock')
    item_list = [item[j:j+1] for j in range(0, len(item),1)]
    for y in item_list:
        pag.keyDown('shift')
        pag.hotkey(y)
        pag.keyUp('shift')

    pag.PAUSE =.2
    pag.press('tab')
    pag.PAUSE =.2
    pag.press('tab')
    pag.PAUSE =.2
    pag.press('enter')
    pag.PAUSE = .25

        # Enter line information
    if item != 'SH':
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite('4162')
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite(str(quantity))
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite(str(price))
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite(str(cost))
        pag.PAUSE =.2
        if item == 'ZFFPKGMAT':
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('space')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('tab')
            pag.PAUSE =.1
            pag.press('enter')
            pag.PAUSE = .5
        else:
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('tab')
            pag.PAUSE = .1
            pag.press('enter')
            pag.PAUSE = .5
    else:
        pag.PAUSE =.2
        pag.typewrite('4162')
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite(str(quantity))
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite(str(price))
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE =.2
        pag.typewrite(str(cost))
        pag.PAUSE =.2
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('tab')
        pag.PAUSE = .1
        pag.press('enter')
        pag.PAUSE = .5

pag.press('tab')
pag.PAUSE = .2
pag.press('tab')
pag.PAUSE = .2
pag.press('tab')
pag.PAUSE = .2
pag.press('enter')


print(ws.max_row)
