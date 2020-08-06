import openpyxl

import pyautogui

import pyperclip

from openpyxl import load_workbook, workbook

wb = load_workbook('production.xlsx')
ws = wb.active
# Office
pyautogui.moveTo(1989,446)
# Home
#pyautogui.moveTo(1249,439)
pyautogui.mouseDown()
pyautogui.PAUSE = .5
pyautogui.mouseUp()
pyautogui.PAUSE = .5


power = 1
for x in ws.rows:
  Order = ws.cell(row = power, column = 1).value
  Amount = ws.cell(row = power, column = 2).value
  power = (power + 1)


  # Order
  pyautogui.click( x = 1715, y = 478 )
  pyautogui.PAUSE = .5
  pyautogui.click( x = 1764, y = 494 )
  pyautogui.PAUSE = .5
  pyautogui.typewrite(str(Order))
  pyautogui.PAUSE = .5
  pyautogui.click( x = 2091, y = 467 )
  pyautogui.PAUSE = .5
  pyautogui.click( x = 2351, y = 507 )
  pyautogui.PAUSE = .5
  pyautogui.click( x = 1810, y = 477 )
  pyautogui.PAUSE = .5
  pyautogui.click( x = 1869, y = 509 )
  pyautogui.PAUSE = .5
  pyautogui.typewrite(str(Amount))
  pyautogui.PAUSE = .5
  pyautogui.click( x = 2229, y = 468 )
  pyautogui.PAUSE = .5


print(ws.max_row)
